#!/usr/bin/env python3
"""Build RepReady.xlsx (one tab per table) from the seed CSVs, for Google Sheets import.

Every cell is written as TEXT (number_format '@') so JSON-string fields, times like
"01:15:00", dates, and ids survive the round-trip exactly. Run after generate_seed.py:
    python build_xlsx.py   ->   RepReady.xlsx
"""
from __future__ import annotations

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

# tab name -> source CSV (order matches the sample-data README)
TABS = [
    ("Athletes", "athletes.csv"),
    ("WorkoutHistory", "workout_history.csv"),
    ("Benchmarks", "benchmarks.csv"),
    ("TrainingRules", "training_rules.csv"),
    ("DemoPrompts", "demo_prompts.csv"),
]

README_ROWS = [
    ["Field", "Value"],
    ["Workbook", "RepReady sample data for Google Sheets import"],
    ["Created for", "HYROX personalized training agent demo"],
    ["Primary tabs", "Athletes, WorkoutHistory, Benchmarks, TrainingRules, DemoPrompts"],
    ["Data format", "List/nested fields are stored as JSON strings so connector tools parse them deterministically."],
    ["Privacy model", "Tools always use a trusted active_user_id and never switch users from chat text."],
    ["Auth model", "Google Sheets/Drive auth is handled by the Claude connector; this workbook contains no credentials."],
    ["History range", "WorkoutHistory fixture data spans 2026-06-01 through 2026-06-13."],
]

HEADER_FILL = PatternFill("solid", fgColor="1F3B5B")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


def _write_sheet(ws, rows: list[list[str]]) -> None:
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.number_format = "@"  # force text — preserves JSON / times / dates verbatim
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    if rows:
        _style_header(ws, len(rows[0]))
        # rough column widths
        for c in range(1, len(rows[0]) + 1):
            width = max((len(str(rows[r][c - 1])) for r in range(min(len(rows), 40)) if c - 1 < len(rows[r])), default=10)
            ws.column_dimensions[get_column_letter(c)].width = min(60, max(12, width + 2))


def main() -> None:
    wb = Workbook()
    # README first
    ws0 = wb.active
    ws0.title = "README"
    _write_sheet(ws0, README_ROWS)

    for tab, fname in TABS:
        path = os.path.join(HERE, fname)
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        ws = wb.create_sheet(title=tab)
        _write_sheet(ws, rows)
        print(f"{tab}: {len(rows) - 1} rows")

    out = os.path.join(HERE, "RepReady.xlsx")
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
